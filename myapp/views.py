import os
from django.conf import settings
from django.http import FileResponse,HttpResponse
from django.shortcuts import render
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime

def index(request):
    return render(request, 'main.html')

def extract_company_name(excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active
    company_name = ws['A1'].value
    return company_name.strip().replace(" ", "_")

def upload_files(request):
    if request.method == 'POST':
        first_excel = request.FILES.get('first_excel')
        second_excel = request.FILES.get('second_excel')

        if first_excel and second_excel:

            current_timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')

            uploads_dir = os.path.join(settings.MEDIA_ROOT, 'uploads')
            os.makedirs(uploads_dir, exist_ok=True)

            second_excel_filename = f"{current_timestamp}_second_excel.xlsx"
            second_excel_path = os.path.join(uploads_dir, second_excel_filename)
            with open(second_excel_path, 'wb+') as destination:
                for chunk in second_excel.chunks():
                    destination.write(chunk)
            
            # Extract company name from the saved second Excel file
            company_name = extract_company_name(second_excel_path)

            # Prepare filenames for the first file and processed files
            first_excel_filename = f"{current_timestamp}_first_excel.xlsx"
            processed_tally_filename = f"{current_timestamp}_{company_name}_processed_tally.xlsx"
            processed_gst_filename = f"{current_timestamp}_{company_name}_processed_gst.xlsx"

            first_excel_path = os.path.join(uploads_dir, first_excel_filename)
            processed_tally_path = os.path.join(uploads_dir, processed_tally_filename)
            processed_gst_path = os.path.join(uploads_dir, processed_gst_filename)

            # Save the first file to disk
            with open(first_excel_path, 'wb+') as destination:
                for chunk in first_excel.chunks():
                    destination.write(chunk)


            # Process files
            process_excels(first_excel_path, second_excel_path, processed_tally_path, processed_gst_path)

            return render(request, 'result.html', {
                'tally_path': processed_tally_path,
                'gst_path': processed_gst_path,
            })
        else:
            return HttpResponse("Please upload both Excel files.")

    return render(request, 'main.html')


def process_excels(GSTR_file_name, Tally_file_name, processed_tally_path, processed_gst_path):
    gst_df = pd.read_excel(GSTR_file_name, sheet_name='B2B', header=5)
    tally_df = pd.read_excel(Tally_file_name, sheet_name='GSTR-3B - Voucher Register', header=8)

    all_tally_df = pd.read_excel(Tally_file_name)
    gst_df = pd.read_excel(GSTR_file_name, sheet_name='B2B', header=5)

    # Function to check if a string matches the date format 'dd-MMM-yy'
    def is_date_format(s):
        try:
            pd.to_datetime(s, format='%d-%b-%y')
            return True
        except ValueError:
            return False

    start_row_id = all_tally_df.index[all_tally_df.iloc[:, 0].apply(is_date_format)].tolist()[0]    


    new_columns_name = all_tally_df.iloc[start_row_id - 1].tolist()

    # print(tally_df.iloc[start_row_id:])
    tally_df = all_tally_df.iloc[start_row_id:].reset_index(drop=True)
    tally_df.columns = new_columns_name

    gst_df = gst_df.iloc[:, [1, 2, 4, 5, 9, 10, 11, 12]]
    gst_df = gst_df.rename(columns={"Unnamed: 1": 'Trade/Legal name', 'Unnamed: 9': 'Taxable Value (₹)'})

    gst_df.columns = gst_df.columns.str.strip()
    tally_df.columns = tally_df.columns.str.strip()

    tally_df = tally_df.dropna(subset=['Vch No.'])
    gst_df = gst_df.dropna(subset=['Invoice number'])

    gst_df = gst_df.loc[:, ~gst_df.columns.duplicated()]
    tally_df = tally_df.loc[:, ~tally_df.columns.duplicated()]

    tally_df['Vch No.'] = tally_df['Vch No.'].astype(str).str.strip()
    gst_df['Invoice number'] = gst_df['Invoice number'].astype(str).str.strip()

    gst_df['GST_sheet_Tax'] = gst_df['Integrated Tax(₹)'] + gst_df['Central Tax(₹)'] + gst_df['State/UT Tax(₹)']
    gst_df['Total_value'] = gst_df['Taxable Value (₹)'] + gst_df['Integrated Tax(₹)'] + gst_df['Central Tax(₹)'] + gst_df['State/UT Tax(₹)']
    tally_df['tally_total_values'] = tally_df['Taxable'] + tally_df['Tax']

    gst_df['gst_Combined_Key'] = gst_df['Invoice number'].str.strip() + '_' + gst_df['Trade/Legal name'].astype(str).str.strip()
    tally_df['tally_Combined_Key'] = tally_df['Vch No.'].str.strip() + '_' + tally_df['Particulars'].astype(str).str.strip()

    grouped_gst_df = gst_df.groupby(['gst_Combined_Key']).agg({
        'Integrated Tax(₹)': 'sum',
        'Central Tax(₹)': 'sum',
        'State/UT Tax(₹)': 'sum',
        'Taxable Value (₹)': 'sum',
        'Total_value': 'sum',
        'GST_sheet_Tax': 'sum',
    }).reset_index()

    grouped_sums = gst_df.groupby('gst_Combined_Key').transform('sum')
    grouped_sums.columns = [f'Sum of {col}' for col in grouped_sums.columns]
    gst_groupby_sum_df = pd.concat([gst_df, grouped_sums], axis=1)

    tally_df['Taxable Amount Difference'] = ''
    tally_df['Tax Amount Difference'] = ''
    tally_df['Total Value Difference'] = ''
    tally_df['Match Status'] = ''

    grouped_gst_df['Taxable Amount Difference'] = ''
    grouped_gst_df['Tax Amount Difference'] = ''
    grouped_gst_df['Total Value Difference'] = ''
    grouped_gst_df['Match Status'] = ''

    for index1, row1 in tally_df.iterrows():
        condition = gst_groupby_sum_df['gst_Combined_Key'] == row1['tally_Combined_Key']
        if condition.any():
            tally_df.at[index1, 'Match Status'] = 'Yes'
            tally_df.at[index1, 'Taxable Amount Difference'] = format(abs(gst_groupby_sum_df[condition].iloc[0]['Sum of Taxable Value (₹)'] - row1['Taxable']), '.2f')
            tally_df.at[index1, 'Tax Amount Difference'] = format(abs(gst_groupby_sum_df[condition].iloc[0]['Sum of GST_sheet_Tax'] - row1['Tax']), '.2f')
            tally_df.at[index1, 'Total Value Difference'] = format(abs(gst_groupby_sum_df[condition].iloc[0]['Total_value'] - row1['tally_total_values']), '.2f')
        else:
            tally_df.at[index1, 'Match Status'] = 'No'

    for index2, row2 in gst_groupby_sum_df.iterrows():
        gst_condition = tally_df['tally_Combined_Key'] == row2['gst_Combined_Key']
        if gst_condition.any():
            matching_row = tally_df.loc[gst_condition].iloc[0]
            gst_groupby_sum_df.at[index2, 'Match Status'] = 'Yes'
            gst_groupby_sum_df.at[index2, 'Taxable Amount Difference'] = format(abs(row2['Sum of Taxable Value (₹)'] - matching_row['Taxable']), '.2f')
            gst_groupby_sum_df.at[index2, 'Tax Amount Difference'] = format(abs(row2['Sum of GST_sheet_Tax'] - matching_row['Tax']), '.2f')
            gst_groupby_sum_df.at[index2, 'Total Value Difference'] = format(abs(row2['Sum of Total_value'] - matching_row['tally_total_values']), '.2f')
        else:
            gst_groupby_sum_df.at[index2, 'Match Status'] = 'No'

    color_tally(tally_df, processed_tally_path)
    color_gst(gst_groupby_sum_df, processed_gst_path)

def color_tally(tally_df, tally_op_path): 
    tally_df.to_excel(tally_op_path, index=False)
    wb = load_workbook(tally_op_path)
    ws = wb.active
    fill_green = PatternFill(start_color="80ff80", end_color="80ff80", fill_type="solid")
    fill_red = PatternFill(start_color="cc0000", end_color="cc0000", fill_type="solid")
    fill_yellow = PatternFill(start_color="ffff1a", end_color="ffff1a", fill_type="solid")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        Match_Status = row[15]
        Taxable_Amount_Difference = row[12]
        Tax_Amount_Difference = row[13]
        Total_Value_Difference = row[14]

        if Match_Status.value == 'No':
            for cell in row:
                cell.fill = fill_red    
        else:
            try:
                # Check if values are numeric and apply conditions
                taxable_diff = float(Taxable_Amount_Difference.value)
                tax_diff = float(Tax_Amount_Difference.value)
                total_diff = float(Total_Value_Difference.value)
                
                if taxable_diff + tax_diff + total_diff > 10:
                    for cell in row:
                        cell.fill = fill_yellow
            except (ValueError, TypeError):
                # Handle non-numeric cases gracefully
                pass  # Continue without filling color if conversion fails

    wb.save(tally_op_path)


def color_gst(gst_df, gst_op_path): 
    # selected_columns = [
    #     'Trade/Legal name', 'Invoice number', 'Invoice Date', 'Invoice Value(₹)', 'Taxable Value (₹)',
    #     'Integrated Tax(₹)', 'Central Tax(₹)', 'State/UT Tax(₹)', 'GST_sheet_Tax', 'Total_value',
    #     'Taxable Amount Difference', 'Tax Amount Difference', 'Total Value Difference', 'Match Status'
    # ]
    # gst_df = gst_df[selected_columns]

    gst_df.to_excel(gst_op_path, index=False)
    wb = load_workbook(gst_op_path)
    ws = wb.active
    fill_green = PatternFill(start_color="80ff80", end_color="80ff80", fill_type="solid")
    fill_red = PatternFill(start_color="cc0000", end_color="cc0000", fill_type="solid")
    fill_yellow = PatternFill(start_color="ffff1a", end_color="ffff1a", fill_type="solid")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        Match_Status = row[-1]
        Taxable_Amount_Difference = row[-4]
        Tax_Amount_Difference = row[-3]
        Total_Value_Difference = row[-2]

        if Match_Status.value == 'No':
            for cell in row:
                cell.fill = fill_red     
        else:
            try:
                # Check if values are numeric and apply conditions
                taxable_diff = float(Taxable_Amount_Difference.value)
                tax_diff = float(Tax_Amount_Difference.value)
                total_diff = float(Total_Value_Difference.value)

                if taxable_diff + tax_diff + total_diff > 10:
                    for cell in row:
                        cell.fill = fill_yellow
            except (ValueError, TypeError):
                pass  # Handle non-numeric cases and continue processing

    wb.save(gst_op_path)


def download_file(request, file_path):
    file_path = os.path.abspath(file_path)
    if os.path.exists(file_path):
        return FileResponse(open(file_path, 'rb'), as_attachment=True)
    else:
        return HttpResponse("File not found.")
