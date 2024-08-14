# import pandas as pd
# import os



# # left == GST
# # right = Tally
# GSTR_file_name = 'Chamunda_gst.xlsx'
# Tally_file_name  = 'GSTR-3B - Voucher Register (4).xlsx'

# gst_df = pd.read_excel(GSTR_file_name, sheet_name='B2B', header=5)
# tally_df = pd.read_excel(Tally_file_name , sheet_name='GSTR-3B - Voucher Register', header=8)

# gst_df = gst_df.iloc[:, [1, 2, 4, 5, 9, 10, 11, 12]]

# # print(gst_df)

# gst_df = gst_df.rename(columns = {"Unnamed: 1":'Trade/Legal name', 'Unnamed: 9':'Taxable Value (₹)'})

# # print("after update colun name ",  tally_df)

# # Strip whitespace from columns
# tally_df.columns = tally_df.columns.str.strip()
# gst_df.columns = gst_df.columns.str.strip()

# # Drop rows with missing key identifiers
# tally_df = tally_df.dropna(subset=['Vch No.'])
# # print("tally_df",tally_df)
# gst_df = gst_df.dropna(subset=['Invoice number'])

# # Remove duplicated columns
# gst_df = gst_df.loc[:, ~gst_df.columns.duplicated()]
# tally_df = tally_df.loc[:, ~tally_df.columns.duplicated()]

# tally_df['Vch No.'] = tally_df['Vch No.'].astype(str).str.strip()
# gst_df['Invoice number'] = gst_df['Invoice number'].astype(str).str.strip()

# gst_df['GST_sheet_Tax'] = gst_df['Integrated Tax(₹)'] + gst_df['Central Tax(₹)'] + gst_df['State/UT Tax(₹)']
# # print("Tally_df with GST_sheet_Tax",tally_df)
# gst_df['Total_value'] = gst_df['Taxable Value (₹)'] + gst_df['Integrated Tax(₹)'] + gst_df['Central Tax(₹)'] + gst_df['State/UT Tax(₹)']
# # print(gst_df)
# tally_df['tally_total_values'] = tally_df['Taxable'] + tally_df['Tax']
# # print("GST_df with tally_total_values",gst_df)

# # Convert relevant columns to integers for comparison
# # gst_df['Taxable Value (₹)'] = gst_df['Taxable Value (₹)'].astype(int)
# # gst_df['Integrated Tax(₹)'] = gst_df['Integrated Tax(₹)'].astype(int)
# # gst_df['Central Tax(₹)'] = gst_df['Central Tax(₹)'].astype(int)
# # gst_df['State/UT Tax(₹)'] = gst_df['State/UT Tax(₹)'].astype(int)
# # gst_df['Total_value'] = gst_df['Total_value'].astype(int)
# # tally_df['Taxable'] = tally_df['Taxable'].astype(int)
# # tally_df['Tax'] = tally_df['Tax'].astype(int)
# # tally_df['tally_total_values'] = tally_df['tally_total_values'].astype(int)

# # Combine 'Invoice number' and 'Trade/Legal name' with trimming
# gst_df['gst_Combined_Key'] = gst_df['Invoice number'].str.strip() + '_' + gst_df['Trade/Legal name'].astype(str).str.strip()
# # print("combine trade name + Invoice number", tally_df)

# # Combine 'Vch No.' and 'Particulars' with trimming
# tally_df['tally_Combined_Key'] = tally_df['Vch No.'].str.strip() + '_' + tally_df['Particulars'].astype(str).str.strip()
# # print("combine trade name + Vch No.", gst_df)


# grouped_gst_df = gst_df.groupby(['gst_Combined_Key']).agg({
#     'Integrated Tax(₹)': 'sum',
#     'Central Tax(₹)': 'sum',
#     'State/UT Tax(₹)': 'sum',
#     'Taxable Value (₹)':'sum',
#     'Total_value':'sum',
#     'GST_sheet_Tax':'sum',

# }).reset_index()
# # print("groupby rows",grouped_tally_df)

# grouped_sums = gst_df.groupby('gst_Combined_Key').transform('sum')

# # Rename columns to indicate they are summed values
# grouped_sums.columns = [f'Sum of {col}' for col in grouped_sums.columns]

# # Combine original DataFrame with the grouped sums
# gst_groupby_sum_df = pd.concat([gst_df, grouped_sums], axis=1)

# # combined_df.to_excel("combined_df.xlsx")
# # print(combined_df)

# difference = []
# missing_in_gst_df = []
# missing_in_tally_df = []

# def is_in_string(substring, string):
#     return substring in string or string in substring

# tally_df['Taxable Amount Difference'] = ''
# tally_df['Tax Amount Difference'] = ''
# tally_df['Total Value Difference'] = ''
# tally_df['Match Status'] = ''

# grouped_gst_df['Taxable Amount Difference'] = ''
# grouped_gst_df['Tax Amount Difference'] = ''
# grouped_gst_df['Total Value Difference'] = ''
# grouped_gst_df['Match Status'] = ''


# # combined_df.to_excel('combined_df.xlsx')

# for index1, row1 in tally_df.iterrows():
#     # print(row1['tally_Combined_Key'])
#     # condition = gst_df['gst_Combined_Key'].str.contains(row1['tally_Combined_Key'])
#     condition = gst_groupby_sum_df['gst_Combined_Key'] == row1['tally_Combined_Key']

#     # row_ids = gst_df[condition].index.tolist()
#     # print("#############",row1['tally_Combined_Key'])
#     # print(gst_df[condition])

#     # print("row_ids===>",row_ids)

#     # print("condition==",condition)
#     if condition.any():
#         tally_df.at[index1, 'Match Status'] = 'Yes'
#         tally_df.at[index1, 'Taxable Amount Difference'] = gst_groupby_sum_df[condition].iloc[0]['Sum of Taxable Value (₹)']-row1['Taxable']
#         tally_df.at[index1, 'Tax Amount Difference'] = gst_groupby_sum_df[condition].iloc[0]['Sum of GST_sheet_Tax']-row1['Tax']
#         tally_df.at[index1, 'Total Value Difference'] = gst_groupby_sum_df[condition].iloc[0]['Total_value']-row1['tally_total_values']

#     else:
#         print("******",row1['tally_Combined_Key'])
#         tally_df.at[index1, 'Match Status'] = 'No'

# tally_df.to_excel("12_07_tally_df.xlsx")

# for index2, row2 in gst_groupby_sum_df.iterrows():
#     gst_condition = tally_df['tally_Combined_Key'] == row2['gst_Combined_Key']

#     if gst_condition.any():
#         matching_row = tally_df.loc[gst_condition].iloc[0]
#         gst_groupby_sum_df.at[index2, 'Match Status'] = 'Yes'
#         gst_groupby_sum_df.at[index2, 'Taxable Amount Difference'] = format(abs(row2['Sum of Taxable Value (₹)'] - matching_row['Taxable']),'.2f')
#         gst_groupby_sum_df.at[index2, 'Tax Amount Difference'] = format(abs(row2['Sum of GST_sheet_Tax'] - matching_row['Tax']),'.2f')
#         gst_groupby_sum_df.at[index2, 'Total Value Difference'] = format(abs(row2['Sum of Total_value'] - matching_row['tally_total_values']),'.2f')
        
#     else:
#         print("gst_missing_key")
#         print("**", row2['gst_Combined_Key'])
#         gst_groupby_sum_df.at[index2, 'Match Status'] = 'No'

# gst_groupby_sum_df.to_excel("15_07_gst_df.xlsx")

import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# left == GST
# right = Tally
GSTR_file_name = '042024_24BYBPG7441P1Z6_GSTR2B_12062024.xlsx'
Tally_file_name  = 'GSTR-3B - Voucher Register (1).xlsx'
# GSTR_file_name = 'Blue_Boy_Gst.xlsx'
# Tally_file_name  = 'Blue_Boy_Tally.xlsx'

# GSTR_file_name = 'Chamunda_gst.xlsx'
# Tally_file_name  = 'Chamunda_tally.xlsx'
# tally_df = pd.read_excel(Tally_file_name , header=8)
# gst_df = pd.read_excel(GSTR_file_name, sheet_name='B2B', header=5)
# tally_df = pd.read_excel(Tally_file_name , sheet_name='GSTR-3B - Voucher Register', header=9)

# tally_df = pd.read_excel(Tally_file_name , header=5)


# GSTR_file_name = 'SQUBEE_Gst.xlsx'
# Tally_file_name  = 'SQUBEE_Tally.xlsx'
tally_df = pd.read_excel(Tally_file_name)
gst_df = pd.read_excel(GSTR_file_name, sheet_name='B2B', header=5)

# print(tally_df)

gst_df = gst_df.iloc[:, [1, 2, 4, 5, 9, 10, 11, 12]]

# print(gst_df)

gst_df = gst_df.rename(columns = {"Unnamed: 1":'Trade/Legal name', 'Unnamed: 9':'Taxable Value (₹)'})

# print("after update colun name ",  tally_df)

# Strip whitespace from columns
tally_df.columns = tally_df.columns.str.strip()
gst_df.columns = gst_df.columns.str.strip()

# Drop rows with missing key identifiers
tally_df = tally_df.dropna(subset=['Vch No.'])
# print("tally_df",tally_df)
gst_df = gst_df.dropna(subset=['Invoice number'])

# Remove duplicated columns
gst_df = gst_df.loc[:, ~gst_df.columns.duplicated()]
tally_df = tally_df.loc[:, ~tally_df.columns.duplicated()]

tally_df['Vch No.'] = tally_df['Vch No.'].astype(str).str.strip()
gst_df['Invoice number'] = gst_df['Invoice number'].astype(str).str.strip()

gst_df['GST_sheet_Tax'] = gst_df['Integrated Tax(₹)'] + gst_df['Central Tax(₹)'] + gst_df['State/UT Tax(₹)']
# print("Tally_df with GST_sheet_Tax",tally_df)
gst_df['Total_value'] = gst_df['Taxable Value (₹)'] + gst_df['Integrated Tax(₹)'] + gst_df['Central Tax(₹)'] + gst_df['State/UT Tax(₹)']
# print(gst_df)
tally_df['tally_total_values'] = tally_df['Taxable'] + tally_df['Tax']
# print("GST_df with tally_total_values",gst_df)

# Convert relevant columns to integers for comparison
# gst_df['Taxable Value (₹)'] = gst_df['Taxable Value (₹)'].astype(int)
# gst_df['Integrated Tax(₹)'] = gst_df['Integrated Tax(₹)'].astype(int)
# gst_df['Central Tax(₹)'] = gst_df['Central Tax(₹)'].astype(int)
# gst_df['State/UT Tax(₹)'] = gst_df['State/UT Tax(₹)'].astype(int)
# gst_df['Total_value'] = gst_df['Total_value'].astype(int)
# tally_df['Taxable'] = tally_df['Taxable'].astype(int)
# tally_df['Tax'] = tally_df['Tax'].astype(int)
# tally_df['tally_total_values'] = tally_df['tally_total_values'].astype(int)

# Combine 'Invoice number' and 'Trade/Legal name' with trimming
gst_df['gst_Combined_Key'] = gst_df['Invoice number'].str.strip() + '_' + gst_df['Trade/Legal name'].astype(str).str.strip()
# print("combine trade name + Invoice number", tally_df)

# Combine 'Vch No.' and 'Particulars' with trimming
tally_df['tally_Combined_Key'] = tally_df['Vch No.'].str.strip() + '_' + tally_df['Particulars'].astype(str).str.strip()
# print("combine trade name + Vch No.", gst_df)


grouped_gst_df = gst_df.groupby(['gst_Combined_Key']).agg({
    'Integrated Tax(₹)': 'sum',
    'Central Tax(₹)': 'sum',
    'State/UT Tax(₹)': 'sum',
    'Taxable Value (₹)':'sum',
    'Total_value':'sum',
    'GST_sheet_Tax':'sum',

}).reset_index()
# print("groupby rows",grouped_tally_df)

grouped_sums = gst_df.groupby('gst_Combined_Key').transform('sum')

# Rename columns to indicate they are summed values
grouped_sums.columns = [f'Sum of {col}' for col in grouped_sums.columns]

# Combine original DataFrame with the grouped sums
gst_groupby_sum_df = pd.concat([gst_df, grouped_sums], axis=1)

# combined_df.to_excel("combined_df.xlsx")
# print(combined_df)
print('-->',"Hyy")
difference = []
missing_in_gst_df = []
missing_in_tally_df = []

def is_in_string(substring, string):
    return substring in string or string in substring

tally_df['Taxable Amount Difference'] = ''
tally_df['Tax Amount Difference'] = ''
tally_df['Total Value Difference'] = ''
tally_df['Match Status'] = ''

gst_groupby_sum_df['Taxable Amount Difference'] = ''
gst_groupby_sum_df['Tax Amount Difference'] = ''
gst_groupby_sum_df['Total Value Difference'] = ''
gst_groupby_sum_df['Match Status'] = ''


# combined_df.to_excel('combined_df.xlsx')

for index1, row1 in tally_df.iterrows():
    # print(row1['tally_Combined_Key'])
    # condition = gst_df['gst_Combined_Key'].str.contains(row1['tally_Combined_Key'])
    condition = gst_groupby_sum_df['gst_Combined_Key'] == row1['tally_Combined_Key']

    # row_ids = gst_df[condition].index.tolist()
    # print("#############",row1['tally_Combined_Key'])
    # print(gst_df[condition])

    # print("row_ids===>",row_ids)

    # print("condition==",condition)
    if condition.any():
        tally_df.at[index1, 'Match Status'] = 'Yes'
        tally_df.at[index1, 'Taxable Amount Difference'] = format(abs(gst_groupby_sum_df[condition].iloc[0]['Sum of Taxable Value (₹)']-row1['Taxable']),'.2f')
        tally_df.at[index1, 'Tax Amount Difference'] = format(abs(gst_groupby_sum_df[condition].iloc[0]['Sum of GST_sheet_Tax']-row1['Tax']),'.2f')
        tally_df.at[index1, 'Total Value Difference'] = format(abs(gst_groupby_sum_df[condition].iloc[0]['Total_value']-row1['tally_total_values']),'.2f')

    else:
        print("******",row1['tally_Combined_Key'])
        tally_df.at[index1, 'Match Status'] = 'No'







for index2, row2 in gst_groupby_sum_df.iterrows():
    gst_condition = tally_df['tally_Combined_Key'] == row2['gst_Combined_Key']

    if gst_condition.any():
        matching_row = tally_df.loc[gst_condition].iloc[0]
        gst_groupby_sum_df.at[index2, 'Match Status'] = 'Yes'
        gst_groupby_sum_df.at[index2, 'Taxable Amount Difference'] = format(abs(row2['Sum of Taxable Value (₹)'] - matching_row['Taxable']),'.2f')
        gst_groupby_sum_df.at[index2, 'Tax Amount Difference'] = format(abs(row2['Sum of GST_sheet_Tax'] - matching_row['Tax']),'.2f')
        gst_groupby_sum_df.at[index2, 'Total Value Difference'] = format(abs(row2['Sum of Total_value'] - matching_row['tally_total_values']),'.2f')
        
    else:
        print("gst_missing_key")
        print("**", row2['gst_Combined_Key'])
        gst_groupby_sum_df.at[index2, 'Match Status'] = 'No'


# tally_df.to_excel("op_Chamunda_tally_df.xlsx")        
# gst_groupby_sum_df.to_excel("op_Chamunda_gst_df.xlsx")

def color_tally(tally_df): 
    tally_op_path = "-op_squbee_tally_df.xlsx"
    tally_df.to_excel(tally_op_path, index=False)
    wb = load_workbook(tally_op_path)
    ws = wb.active
    fill_green = PatternFill(start_color="80ff80", end_color="80ff80", fill_type="solid")
    fill_red = PatternFill(start_color="ff8566", end_color="ff8566", fill_type="solid")
     #ff8566

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        match_status = row[15]
        if match_status.value =='No':
            for cell in row:
                cell.fill = fill_red     

    wb.save(tally_op_path)   
tally_df = tally_df[['Date', 'Particulars',	'Vch Type',	'Vch No.', 'Taxable', 'IGST', 'CGST', 'SGST/', 'Cess', 'Tax', 'Amount_of_tax', 'tally_Combined Key', 'Taxable Amount Difference', 'Tax Amount Difference', 'Invoice Value Difference', 'tally_total_values', 'tally_Combined_Key', 'Total Value Difference', 'Match Status']]    

color_tally(tally_df)



def color_gst(gst_df): 
    gst_op_path = "-op_squbee_gst_df.xlsx"
    gst_df.to_excel(gst_op_path, index=False)
    wb = load_workbook(gst_op_path)
    ws = wb.active
    fill_green = PatternFill(start_color="80ff80", end_color="80ff80", fill_type="solid")
    fill_red = PatternFill(start_color="ff8566", end_color="ff8566", fill_type="solid")
     #ff8566

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        match_status = row[13]
        if match_status.value=='No':
            for cell in row:
                cell.fill = fill_red     

    wb.save(gst_op_path)       
gst_df = gst_groupby_sum_df[['Trade/Legal name', 'Invoice number', 'Invoice Date', 'Invoice Value(₹)', 'Taxable Value (₹)',	'Integrated Tax(₹)', 'Central Tax(₹)', 'State/UT Tax(₹)', 'GST_sheet_Tax', 'Total_value', 'Taxable Amount Difference', 'Tax Amount Difference', 'Total Value Difference', 'Match Status']]


color_gst(gst_df)

# gst_op_path = "op_squbee_gst_df.xlsx"

# tally_df.to_excel("")
# gst_groupby_sum_df.to_excel("op_squbee_gst_df.xlsx")